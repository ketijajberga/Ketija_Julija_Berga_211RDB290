"""
VPN veiktspējas mērījumu statistikas aprēķins
Ievaddati:  VPN_Neapstradatie_merijumi.xlsx
Izvaddati:  VPN_Statistika_rezultati.xlsx
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# KONFIGURĀCIJA
# ============================================================

INPUT_FILE  = "VPN_Neapstradatie_merijumi.xlsx"
OUTPUT_FILE = "VPN_Statistika_rezultati.xlsx"

SOLUTIONS = ["IPsec", "Netmaker", "OpenVPN", "WireGuard", "ZeroTier"]

PROFILES_ORDER = [
    "Joslas platums 100 Mbit/s",
    "Joslas platums 50 Mbit/s",
    "Joslas platums 10 Mbit/s",
    "Svārstīgums 100ms ±30ms",
    "Svārstīgums 100ms ±80ms",
    "Aizture 50ms",
    "Aizture 100ms",
    "Aizture 200ms",
    "Paketes zudums 0.5%",
    "Paketes zudums 1%",
]

METRICS = {
    "Caurlaidspēja (Mbps)": "Caurlaidspēja (Mbps)",
    "Aizture vid. (ms)":    "Aizture (ms)",
    "CPU (%)":              "CPU (%)",
    "Paketes zudums (%)":   "Paketes zudums (%)",
}

STAT_LABELS = ["Vid.", "Med.", "St.nov.", "Min.", "Maks.", "P95"]

SCENARIOS = [
    ("LAN-LAN",             "LAN-LAN"),
    ("Attālinātā piekļuve", "Attālinātā piekļuve"),
]

# ============================================================
# 1. DATU IELĀDE — apvieno visas lapas vienā DataFrame
# ============================================================

print("Lasu datus no:", INPUT_FILE)
all_sheets = pd.read_excel(INPUT_FILE, sheet_name=None)

frames = []
for sheet_name, df in all_sheets.items():
    frames.append(df)
raw = pd.concat(frames, ignore_index=True)

print(f"  Kopā rindas: {len(raw)}")

# ============================================================
# 2. STATISTIKAS APRĒĶINS
# ============================================================

def calc_stats(series):
    """
    Aprēķina 6 statistikas rādītājus dotajai skaitliskai sērijai.
    Atgriež [vidējā, mediāna, std, min, max, p95] vai [None]*6 ja nav datu.
    """
    s = pd.to_numeric(series, errors="coerce").dropna()
    if len(s) == 0:
        return [None] * 6
    return [
        round(s.mean(),         3),  # Vidējā vērtība
        round(s.median(),       3),  # Mediāna
        round(s.std(),          3),  # Standartnovirze
        round(s.min(),          3),  # Minimums
        round(s.max(),          3),  # Maksimums
        round(s.quantile(0.95), 3),  # 95. procentile
    ]

# Aprēķina statistiku katram risinājumam / scenārijam / profilam / metrikai
results = {}
for sol in SOLUTIONS:
    results[sol] = {}
    for scenario_key, scenario_label in SCENARIOS:
        results[sol][scenario_key] = {}

        # Filtrē pēc risinājuma un scenārija
        mask = (raw["Risinājums"] == sol) & (raw["Scenārijs"] == scenario_label)
        df_sc = raw[mask]

        for profile in PROFILES_ORDER:
            df_pr = df_sc[df_sc["Profils"] == profile]
            results[sol][scenario_key][profile] = {}

            for col_name in METRICS.keys():
                if col_name in df_pr.columns:
                    results[sol][scenario_key][profile][col_name] = calc_stats(df_pr[col_name])
                else:
                    results[sol][scenario_key][profile][col_name] = [None] * 6

print("Statistika aprēķināta.")

# ============================================================
# 3. EXCEL IZVADE
# ============================================================

H_FILL  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
SH_FILL = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
ALT     = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
H_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
SH_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=9)
N_FONT  = Font(name="Arial", size=9)
B_FONT  = Font(name="Arial", bold=True, size=9)
CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT    = Alignment(horizontal="left",   vertical="center")
THIN    = Side(style="thin", color="AAAAAA")
BRD     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()
wb.remove(wb.active)

for scenario_key, scenario_label in SCENARIOS:
    ws = wb.create_sheet(title=scenario_label)
    ws.freeze_panes = "B5"

    total_cols = 1 + len(SOLUTIONS) * len(METRICS) * len(STAT_LABELS)

    # 1. rinda — virsraksts
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws["A1"] = f"Veiktspēja pasliktinātos tīkla apstākļos — {scenario_label} scenārijs"
    ws["A1"].font = Font(name="Arial", bold=True, size=12, color="1F4E79")
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 22

    # 2. rinda — risinājumu virsraksti
    ws.cell(row=2, column=1, value="Degradācijas profils").font = SH_FONT
    ws.cell(row=2, column=1).fill = SH_FILL
    ws.cell(row=2, column=1).alignment = CENTER
    ws.cell(row=2, column=1).border = BRD
    ws.row_dimensions[2].height = 20

    col = 2
    for sol in SOLUTIONS:
        span = len(METRICS) * len(STAT_LABELS)
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + span - 1)
        c = ws.cell(row=2, column=col, value=sol)
        c.font = H_FONT; c.fill = H_FILL; c.alignment = CENTER; c.border = BRD
        col += span

    # 3. rinda — metriku virsraksti
    ws.cell(row=3, column=1).border = BRD
    ws.row_dimensions[3].height = 20

    col = 2
    for sol in SOLUTIONS:
        for col_name, metric_label in METRICS.items():
            ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + len(STAT_LABELS) - 1)
            c = ws.cell(row=3, column=col, value=metric_label)
            c.font = SH_FONT; c.fill = SH_FILL; c.alignment = CENTER; c.border = BRD
            col += len(STAT_LABELS)

    # 4. rinda — statistikas apzīmējumi
    ws.cell(row=4, column=1).border = BRD
    ws.row_dimensions[4].height = 16

    col = 2
    for sol in SOLUTIONS:
        for _ in METRICS:
            for stat in STAT_LABELS:
                c = ws.cell(row=4, column=col, value=stat)
                c.font = SH_FONT; c.fill = SH_FILL; c.alignment = CENTER; c.border = BRD
                col += 1

    # Datu rindas
    for i, profile in enumerate(PROFILES_ORDER):
        row = 5 + i
        ws.row_dimensions[row].height = 16
        fill = ALT if i % 2 == 0 else PatternFill()

        c = ws.cell(row=row, column=1, value=profile)
        c.font = B_FONT; c.alignment = LEFT; c.border = BRD
        if i % 2 == 0:
            c.fill = fill

        col = 2
        for sol in SOLUTIONS:
            for col_name in METRICS.keys():
                vals = results[sol][scenario_key][profile][col_name]
                for val in vals:
                    c = ws.cell(row=row, column=col)
                    c.value = val if val is not None else "N/A"
                    c.font = N_FONT; c.alignment = CENTER; c.border = BRD
                    if i % 2 == 0:
                        c.fill = fill
                    col += 1

    # Kolonnu platumi
    ws.column_dimensions["A"].width = 26
    for col_idx in range(2, col):
        ws.column_dimensions[get_column_letter(col_idx)].width = 8

wb.save(OUTPUT_FILE)
print(f"Rezultāti saglabāti: {OUTPUT_FILE}")
